"""XLSX renderer — one sheet per detected table + a "Text" sheet for everything else.

Preserves merged cells via `openpyxl` merge ranges. Non-table blocks go to a
"Text" sheet (block_index, page, kind, text, corrected_text) so nothing is
lost at the format boundary.
"""
from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from ocr_agentic_engine.renderers._ordering import iter_blocks
from ocr_agentic_engine.types import Annotations, DocumentRepresentation, OCRBlock


def render_xlsx_tabular(doc: DocumentRepresentation, ann: Annotations) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    table_idx = 0
    for idx, b in iter_blocks(doc, ann):
        if b.kind == "table" and b.table_cells:
            table_idx += 1
            _write_table_sheet(wb, f"Table{table_idx}_p{b.page}", b)

    _write_text_sheet(wb, doc, ann)

    buf = BytesIO(); wb.save(buf); return buf.getvalue()


def _write_table_sheet(wb: Workbook, title: str, b: OCRBlock) -> None:
    ws = wb.create_sheet(title[:31])
    for c in b.table_cells or []:
        ws.cell(row=c.row + 1, column=c.col + 1, value=c.text)
        if c.row_span > 1 or c.col_span > 1:
            ws.merge_cells(start_row=c.row + 1, start_column=c.col + 1,
                            end_row=c.row + c.row_span,
                            end_column=c.col + c.col_span)
    # widen columns
    if b.table_cells:
        max_col = max(c.col + c.col_span for c in b.table_cells)
        for i in range(max_col):
            ws.column_dimensions[get_column_letter(i + 1)].width = 18


def _write_text_sheet(wb: Workbook, doc: DocumentRepresentation, ann: Annotations) -> None:
    ws = wb.create_sheet("Text")
    ws.append(["block_index", "page", "kind", "text", "corrected_text", "heading_level"])
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for idx, b in iter_blocks(doc, ann):
        ws.append([
            idx, b.page, b.kind, b.text,
            ann.corrections.get(idx, ""),
            ann.headings.get(idx, ""),
        ])
    ws.column_dimensions["D"].width = 60
    ws.column_dimensions["E"].width = 60
